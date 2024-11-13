import math
import csv
import os

from django.shortcuts import render, get_list_or_404, get_object_or_404
from django.utils import timezone
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from django.db.models import Sum
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from xhtml2pdf import pisa
import pandas as pd

#-- Modelos requeridos
from apps.app_id_liquidacion.models import IdLiquidacion
from apps.app_beneficiarios.models import Beneficiario
from apps.app_codigos.models import Codigo
from apps.app_ubicacion.models import Ubicacion
from apps.app_movimiento_mensual.models import MovimientoMensual
from .models import DetalleLiquidacion
from .forms import BusquedaReciboLiquidacionForm


@login_required
def generar_detalle_liquidacion(request, pk=None):

    context = {
        'object_list' : IdLiquidacion.objects.filter(cerrado=False),
        'titulo': 'Procesar Liquidación',
        'subtitulo': 'Generar liquidación',
        'errores': []       
    }
    
    if pk is not None:

        try:
            # Obtengo la liquidación activa
            liquidacion_actual = IdLiquidacion.objects.filter(pk=pk, cerrado=False).first()
            # Si había una versión anterior, la elimino
            if liquidacion_actual:
                DetalleLiquidacion.objects.filter(liquidacion=liquidacion_actual).delete()
            # Obtengo el listado de beneficiarios activos
            beneficiarios = Beneficiario.objects.filter(activo=True)

        except Exception as e:
            context['errores'].append(F'Error -> {e} ')

        else:        
            # Auxiliar para recibos
            recibo = 0
            # --- Proceso de liquidación
            for b in beneficiarios:
                try:                   
                    # Obtengo los códigos activos para ese tipo de liquidación
                    codigos_a_liquidar = Codigo.objects.filter(pension=b.tipo_pension, activo=True)

                    # Variables totalizadoras
                    total_remunerativos = 0
                    total_no_remunerativos = 0
                    total_descuentos = 0
                    # Número de recibo
                    recibo += 1

                    for c in codigos_a_liquidar:
                        movmes = MovimientoMensual.objects.filter(beneficiario=b, codigo=c, id_liquidacion=liquidacion_actual, pendiente=True).first()
                        temporal_valor = 0

                        # Si no hay registros en el movmes, se calculan según el tipo de cálculo establecido en el código
                        if movmes is None:            

                            #--- Cálculo del valor
                            if c.calculo == Codigo.TipoCalculo.VALOR_FIJO:
                                temporal_valor = c.valor

                            if c.calculo == Codigo.TipoCalculo.PORC_SOBRE_REMUNERATIVO:
                                temporal_valor = total_remunerativos * c.valor 

                            if c.calculo == Codigo.TipoCalculo.PORC_SOBRE_TOTAL_PAGO:
                                temporal_valor = (total_remunerativos + total_no_remunerativos) * c.valor

                            if c.tipo == Codigo.TipoConcepto.REDONDEO_POSITIVO:
                                temporal_valor = float(math.ceil(total_remunerativos + total_no_remunerativos)) - (total_remunerativos + total_no_remunerativos)

                            if c.tipo == Codigo.TipoConcepto.REDONDEO_NEGATIVO:
                                temporal_valor = float(math.ceil(total_descuentos)) - (total_descuentos)

                        else:
                            temporal_valor = movmes.importe

                        #--- Acumular de acuerdo al concepto
                        if c.tipo == Codigo.TipoConcepto.REMUNERATIVO or c.tipo == Codigo.TipoConcepto.REDONDEO_POSITIVO: total_remunerativos += temporal_valor
                        if c.tipo == Codigo.TipoConcepto.NO_REMUNERATIVO: total_no_remunerativos +=  temporal_valor
                        if c.tipo == Codigo.TipoConcepto.DESCUENTO or c.tipo == Codigo.TipoConcepto.REDONDEO_NEGATIVO: total_descuentos += temporal_valor 

                        if c.signo == Codigo.SignoImporte.NEGATIVO:
                            temporal_valor = - temporal_valor

                        if temporal_valor != 0:
                            #--- Agrego el objeto a la bbdd, si el valor no es cero
                            detalle = DetalleLiquidacion(liquidacion=liquidacion_actual, 
                                                            recibo=recibo, 
                                                            localidad=b.ubicacion, 
                                                            legajo=b.legajo, 
                                                            beneficiario=b.apellido_y_nombre,
                                                            documento=b.documento, 
                                                            cuil=b.cuil,
                                                            tipo_pension=b.tipo_pension.codigo_numerico, 
                                                            codigo=c.codigo_numerico, 
                                                            codigo_descripcion=c.descripcion, 
                                                            codigo_tipo = Codigo.TipoConcepto(c.tipo).label, 
                                                            importe=temporal_valor)
                            detalle.save()

                except Exception as e:
                    context['errores'].append(F'Error Beneficiario {b} Código {c}')
                    print('Excepcion interna ', e)
                    continue


            liquidacion_actual.fecha_ultima_ejecucion = timezone.now()
            liquidacion_actual.save()

            context['liquidacion_ejecutada'] = liquidacion_actual.__str__()
            context['beneficiarios_count'] = beneficiarios.count()
            context['registros_count'] = DetalleLiquidacion.objects.filter(liquidacion=liquidacion_actual).count()

    return render(request, 'detalleliqs/ejecutar_liquidacion.html', context)
    



def generar_reporte_csv_liquidacion(request, idliq):
    """
    Genera un archivo CSV con el detalle de la liquidación
    Actualmente, Vista sin mapeo de URL 
    """

    liquidacion_actual = IdLiquidacion.objects.filter(pk=idliq, cerrado=False).first()
    detalle_liquidacion = get_list_or_404(DetalleLiquidacion, liquidacion=liquidacion_actual)

    titulo = F"reporte_{liquidacion_actual}.csv"

    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": F'attachment; filename="{titulo}"'},)
    
    response.encoding = 'UTF-8'
    
    writer = csv.writer(response, delimiter=';')

    # Títulos de las columnas
    titulos = ['ID_LIQUIDACIÓN', 'DESC_LIQUIDACIÓN', 'FPAGO_LIQUIDACIÓN', 'FEJEC_LIQUIDACIÓN', 'RECIBO', 'LOCALIDAD',
               'ID_LOCALIDAD', 'LEGAJO', 'BENEFICIARIO', 'DOCUMENTO', 'CUIL', 'TIPO_PENSIÓN', 'CONCEPTO', 'DESC_CONCEPTO', 'IMPORTE']
    writer.writerow(titulos)

    registros = []

    for detalle in detalle_liquidacion:
        registros.append([detalle.liquidacion.id, detalle.liquidacion, detalle.liquidacion.fecha_pago, detalle.liquidacion.fecha_ultima_ejecucion,
                         detalle.recibo, detalle.localidad, detalle.localidad.id, detalle.legajo, detalle.beneficiario, detalle.documento,
                          detalle.cuil, detalle.tipo_pension, detalle.codigo, detalle.codigo_descripcion, detalle.importe ])
        
    writer.writerows(registros)

    return response
    

def auxiliar_generar_consulta(idliq : int ) -> Workbook:
    """
    Función auxiliar para generar los datos en formato xlsx
    """
    liquidacion_actual = IdLiquidacion.objects.filter(pk=idliq, cerrado=False).first()
    detalle_liquidacion = get_list_or_404(DetalleLiquidacion, liquidacion=liquidacion_actual)

    return detalle_liquidacion


@login_required
def generar_reporte_xlsx_liquidacion(request, idliq):
    """
    Genera reporte XLSX con el detalle de la liquidación 
    La primera hoja contiene el listado desagregado de todos los beneficiarios
    liquidados y sus correspondientes conceptos
    La segunda muestra un resumen agrupado por localidad de los totales liquidados
    por concepto    
    """

    detalle_liquidacion = auxiliar_generar_consulta(idliq)

    titulos = ['ID_LIQUIDACIÓN', 'DESC_LIQUIDACIÓN', 'FPAGO_LIQUIDACIÓN', 'FEJEC_LIQUIDACIÓN', 'RECIBO', 'LOCALIDAD',
               'ID_LOCALIDAD', 'LEGAJO', 'BENEFICIARIO', 'DOCUMENTO', 'CUIL', 'TIPO_PENSIÓN', 'CONCEPTO', 'DESC_CONCEPTO', 'IMPORTE']

    wb = Workbook()
    wb.remove_sheet(wb.active)
    ws = wb.create_sheet('Reporte Vertical')

    # Títulos de la hoja de reporte vertical
    c = 1
    for t in titulos:
        ws.cell(row=1, column=c, value=t)
        c += 1

    # Datos de la hoja de reporte vertical
    f = 2
    for detalle in detalle_liquidacion:
        ws.cell(row=f, column=1, value=detalle.liquidacion.id)
        ws.cell(row=f, column=2, value=detalle.liquidacion.__str__())
        ws.cell(row=f, column=3, value=str(detalle.liquidacion.fecha_pago))
        ws.cell(row=f, column=4, value=str(detalle.liquidacion.fecha_ultima_ejecucion))
        ws.cell(row=f, column=5, value=detalle.recibo)
        ws.cell(row=f, column=6, value=detalle.localidad.nombre)
        ws.cell(row=f, column=7, value=detalle.localidad.id)
        ws.cell(row=f, column=8, value=detalle.legajo)
        ws.cell(row=f, column=9, value=detalle.beneficiario)
        ws.cell(row=f, column=10, value=detalle.documento)
        ws.cell(row=f, column=11, value=detalle.cuil)
        ws.cell(row=f, column=12, value=detalle.tipo_pension)
        ws.cell(row=f, column=13, value=detalle.codigo)
        ws.cell(row=f, column=14, value=detalle.codigo_descripcion)
        ws.cell(row=f, column=15, value=detalle.importe)
        f += 1

    # Generar reporte agrupado
    df = pd.DataFrame(ws.values, columns=titulos)
    df = df[['TIPO_PENSIÓN', 'LOCALIDAD', 'CONCEPTO', 'IMPORTE']]

    df1 = df.pivot_table(index=['TIPO_PENSIÓN', 'LOCALIDAD'], columns='CONCEPTO', values='IMPORTE', aggfunc='sum')

    ws_agrupado = wb.create_sheet('Reporte Agrupado')

    for r in dataframe_to_rows(df1, index=True, header=True):
        ws_agrupado.append(r)

    titulo = "reporte_liquidacion.xlsx"
    response = HttpResponse(
        content_type="application/ms-excel",
        headers={"Content-Disposition": F'attachment; filename="{titulo}"'},)

    wb.save(response) 
    return response
        
@login_required
def generar_recibo_pdf(request, idliq):
    """
    Vista que genera recibos en PDF para liquidación cerrada
    """
    liquidacion = IdLiquidacion.objects.get(id=idliq, cerrado=True)
    resultados = get_list_or_404(DetalleLiquidacion, liquidacion=liquidacion)

    context = {'detalle_list' : resultados,
               'liquidacion': liquidacion }

    template_path = 'detalleliqs/recibos.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="recibos.pdf"'

    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)

    return response


@login_required
def simulador_recibo(request):
    """
    Consulta de detalles de conceptos en una liquidación en particular
    Solamente muestra los resultados en liquidaciones abiertas
    """

    form = BusquedaReciboLiquidacionForm()
    mensaje = ''

    if request.method == 'POST':
        form = BusquedaReciboLiquidacionForm(request.POST)

        if form.is_valid():
            cd = form.cleaned_data

            try:
                idliquidacion = IdLiquidacion.objects.get(id=cd['idliq'].id)
                idlocalidad = Ubicacion.objects.get(id=cd['idubicacion'].id)

                object_list = DetalleLiquidacion.objects.filter(liquidacion=idliquidacion, localidad=idlocalidad, legajo=cd['idlegajo'])

                if len(object_list) == 0:
                    raise Exception("No hay resultados para mostrar") 

                return render(request, 'detalleliqs/simulador_recibos.html', { 
                                                                                'titulo': 'SIM RECIBOS',
                                                                                'subtitulo': 'Simulador de RECIBOS - RESULTADOS',
                                                                                'liquidacion': idliquidacion,
                                                                                'object_list': object_list})
            except Exception as e:
                return render(request, 'buscar.html', {
                                                'titulo': 'SIM RECIBOS',
                                                'subtitulo': F'Simulador de RECIBOS | ERRORES: {e}',
                                                'form': form,
                                                'template_tabla': 'detalleliqs/tabla.html'
                                                  })

    else:
        return render(request, 'buscar.html', {
                                                'titulo': 'SIM RECIBOS',
                                                'subtitulo': 'Simulador de RECIBOS',
                                                'form': form,
                                                'mensaje': mensaje,
                                                'template_tabla': 'detalleliqs/tabla.html'
                                                  })