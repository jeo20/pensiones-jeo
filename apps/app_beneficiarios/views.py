from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.views.generic import DetailView, ListView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from .models import Beneficiario
from .forms import BusquedaBeneficiario, ReporteBeneficiariosForm

# Create your views here.

class BeneficiarioListView(LoginRequiredMixin, ListView):
    model = Beneficiario
    fields = '__all__'
    template_name = 'lista.html'
    extra_context = {'titulo': 'Beneficiario', 'subtitulo': 'Listado de Beneficiarios', 'template_tabla': 'beneficiarios/tabla.html'}
    paginate_by = 12

class BeneficiarioCreateView(LoginRequiredMixin,CreateView):
    model = Beneficiario
    fields = ['ubicacion', 'legajo', 'apellido_y_nombre', 'genero', 'fecha_nacimiento', 'imagen', 'documento_tipo', 'documento',  'cuil',
              'direccion', 'tipo_pension', 'cobra_sac', 'activo']
    extra_context = { 'titulo': 'Beneficiario', 'subtitulo': 'Crear nuevo Beneficiario'}
    template_name = 'crear.html'

    def form_valid(self, form):
        form.instance.creado_usuario = self.request.user.username
        return super().form_valid(form)

class BeneficiarioDetailView(LoginRequiredMixin, DetailView):
    model = Beneficiario
    fields = '__all__'
    extra_context = { 'titulo':'Beneficiario', 'subtitulo': 'Actualizar datos'}
    template_name = 'beneficiarios/detalle.html'

class BeneficiarioUpdateView(LoginRequiredMixin, UpdateView):
    model = Beneficiario
    fields = ['ubicacion', 'legajo', 'apellido_y_nombre', 'genero', 'fecha_nacimiento', 'imagen', 'documento_tipo', 'documento',  'cuil',
              'direccion', 'tipo_pension', 'cobra_sac', 'activo']
    extra_context = { 'titulo': 'Beneficiario', 'subtitulo': 'Actualizar Beneficiario'}
    template_name = 'crear.html'

    def form_valid(self, form):
        form.instance.modificado_usuario = self.request.user.username
        return super().form_valid(form)


@login_required
def buscar_beneficiarios(request):
    """
    Búsqueda de beneficiarios por diferentes criterios
    """
    form = BusquedaBeneficiario()
    mensaje = "Ingrese los criterios de búsqueda"
    object_list = None 

    if request.method == 'POST':
        form = BusquedaBeneficiario(request.POST)

        if form.is_valid():
            cd = form.cleaned_data
            criterio = cd['criterio']
            campo = cd['campo']
            object_list = []

            if criterio.isnumeric():
                if campo == 'DNI':
                    object_list = Beneficiario.objects.filter(documento=criterio)

                if campo == 'LEG':
                    object_list = Beneficiario.objects.filter(legajo=criterio)
            else:
                if campo == 'APE':
                    object_list = Beneficiario.objects.filter(apellido_y_nombre__contains=criterio)

            if len(object_list) == 0:
                mensaje = 'No se encontraron resultados con los criterios de búsqueda elegidos'
            
    return render(request, 'buscar.html', {'form': form,
                                            'mensaje': mensaje,
                                            'titulo': 'Beneficiarios',
                                            'subtitulo': 'Buscar Beneficiarios',
                                            'object_list': object_list,
                                            'template_tabla': 'beneficiarios/tabla.html'})


@login_required
def generar_reportes_xlsx(request):
    """
    Vista con búsqueda multiparamétrica que genera reportes en Excel
    con información sobre los BENEFICIARIOS
    Campos de búsqueda:
        * Ubicación
        * Tipo Pensión
        * Género
        * Estado
        * SAC
    """

    context = {
        'form' : ReporteBeneficiariosForm(),
        'titulo': 'Beneficiarios',
        'subtitulo': 'Generación de Reportes',
        'errores': []
    }

    TITULOS = ['UBICACIÓN DESCRIPCION', 'UBICACION', 'LEGAJO', 'APELLIDO Y NOMBRE', 'FECHA NACIMIENTO',
               'GENERO', 'GENERO DESCRIPCÓN',
               'DOCUMENTO TIPO', 'DOCUMENTO TIPO DESC', 'DOCUMENTO NUMERO',
               'CUIL', 'DIRECCIÓN', 
               'TIPO PENSIÓN', 'TIPO PENSION DESC',
               'COBRA SAC', 'ACTIVO', 'CREADO FECHA', 'ÚLTIMA MODIFICACIÓN']
    
    if request.method == 'POST':
        form = ReporteBeneficiariosForm(request.POST)

        print(form.is_valid())
        print(form.cleaned_data)

        if form.is_valid():
            cd = form.cleaned_data 
            kwargs = {}

            if cd['ubicacion']:
                kwargs['ubicacion__in'] = cd['ubicacion']
            
            if cd['tipo_pension']:
                kwargs['tipo_pension__in'] = cd['tipo_pension']

            if cd['genero']:
                kwargs['genero'] = Beneficiario.Genero(cd['genero'])

            if cd['estado']:
                kwargs['activo'] = True if cd['estado'] == 'T' else False

            if cd['sac']:
                kwargs['cobra_sac'] = True if cd['sac'] == 'T' else False

            try:
                object_list = Beneficiario.objects.filter(**kwargs)

                if len(object_list) == 0:
                    raise Exception('No hay resultados para esos criterios de búsqueda')
                
                wb = Workbook()
                ws = wb.active
                ws.title = cd['nombre_reporte'].upper()

                for idx, t in enumerate(TITULOS):
                    ws.cell(row=1, column=idx+1, value=t)

                fila = 2
                for beneficiario in object_list:
                    ws.cell(row=fila, column=1, value=beneficiario.ubicacion.__str__())
                    ws.cell(row=fila, column=2, value=beneficiario.ubicacion.localidad)
                    ws.cell(row=fila, column=3, value=beneficiario.legajo)
                    ws.cell(row=fila, column=4, value=beneficiario.apellido_y_nombre)
                    ws.cell(row=fila, column=5, value=beneficiario.fecha_nacimiento)
                    ws.cell(row=fila, column=6, value=beneficiario.genero)
                    ws.cell(row=fila, column=7, value=Beneficiario.Genero(beneficiario.genero).name)
                    ws.cell(row=fila, column=8, value=beneficiario.documento_tipo)
                    ws.cell(row=fila, column=9, value=Beneficiario.TipoDocumento(beneficiario.documento_tipo).name)
                    ws.cell(row=fila, column=10, value=beneficiario.documento)
                    ws.cell(row=fila, column=11, value=beneficiario.cuil)
                    ws.cell(row=fila, column=12, value=beneficiario.direccion)
                    ws.cell(row=fila, column=13, value=beneficiario.tipo_pension.codigo_numerico)
                    ws.cell(row=fila, column=14, value=beneficiario.tipo_pension.descripcion_corta)
                    ws.cell(row=fila, column=15, value=beneficiario.cobra_sac)
                    ws.cell(row=fila, column=16, value=beneficiario.activo)
                    ws.cell(row=fila, column=17, value=beneficiario.creado.__str__())
                    ws.cell(row=fila, column=18, value=beneficiario.modificado.__str__())
                    fila += 1
            
                titulo = F"{cd['nombre_reporte']}.xlsx"
                response = HttpResponse(content_type="application/ms-excel",
                                    headers={"Content-Disposition": F'attachment; filename="{titulo}"'},)
                
                wb.save(response) 
                return response
            
            except Exception as e:
                context['errores'].append(F"Ha ocurrido un error -> {e}")

    return render(request, 'beneficiarios/reportes.html', context)